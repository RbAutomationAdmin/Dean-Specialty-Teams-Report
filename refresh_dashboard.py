"""Refresh the Specialty Teams Dashboard from Qlik exports.

Workflow:
  1. In Qlik, open the appointment-detail table (23 columns: CalendarDate, ApptSetBy,
     Ownership Group, StoreName, ..., AppointmentSetOn, '# of attempts', Source),
     filtered to the specialty-team reps.
  2. Export it. Split the export (by year, by team - any chunking works) if it hits
     Qlik's export limit. Files land in Downloads as "Untitled - <date>.xlsx" -
     no renaming needed.
  3. Run this script (or double-click "Refresh Dashboard.bat").

The script finds today's exports in Downloads (or takes explicit --files), rebuilds
the MASTER data blob embedded in specialty_teams_dashboard.html, and stamps the
refresh time. The roster (team assignments + tenure windows) lives in roster.csv
next to this script - edit that file to add/move/retire agents.

Rules mirrored from the dashboard's Methodology card:
  - A row is attributed to a team by rep name + tenure window (the '@XXXX - '
    prefix on ApptSetBy is ignored). Rows outside every tenure window are dropped.
  - All metrics are gated by CalendarDate at view time; AppointmentSetOn is carried
    through but not used as a filter.

Overlap handling: exports often overlap (re-downloads, per-team chunks, partial date
windows). For every (calendar day, rep+team) the NEWEST export that covers that rep
and team for that day wins; older files only fill gaps. So a fresh full-history
export supersedes everything, while per-team or per-period chunks merge cleanly.

Carry-forward: the dashboard's current embedded data participates in the merge as the
OLDEST source. Because the Qlik data model is dynamic (no history - a rep who leaves a
specialty team gets re-prefixed, e.g. back to @OUTB, and falls out of prefix-filtered
exports), a full-YTD export each refresh is enough: fresh exports supersede every
(day, rep+team) they cover, while prior-year history and former reps' records persist
from the dashboard itself. Use --no-carry to rebuild strictly from the given exports.
"""
import argparse
import csv
import datetime
import glob
import json
import os
import re
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(HERE, "specialty_teams_dashboard.html")
ROSTER_PATH = os.path.join(HERE, "roster.csv")
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")

BASE = datetime.date(2025, 1, 1)  # must match MASTER.base in the HTML
METRICS = ["attempts", "contacts", "apptsset", "sameday", "nextday", "issued",
           "demos", "nodemo", "nothome", "netsales", "netsale$"]
N_ACT = 2

# Export columns required, in the order the metrics are read.
COLS = ["CalendarDate", "ApptSetBy", "StoreName", "# of attempts", "# of Contacts",
        "Appts Set", "Appts Set Same Day", "Appts Set Next Day", "Appts Issued",
        "Demos", "No Demo", "Not Home", "Net Sales", "Net Sale $", "AppointmentSetOn"]

# Cost-center store export (Agent Overview (Store) sheet, Store tab, NO agent filter,
# YTD date range): store-level totals for the ENTIRE cost center, used by the
# "YTD vs Entire Cost Center" card. (Column name, CC metric) in embed order.
CC_COLMAP = [("# of Call Attempts", "attempts"), ("# of Contacts", "contacts"),
             ("Appts Set", "apptsset"), ("Appts Issued", "issued"),
             ("Net Issued Appts", "netissued"), ("Net Orders", "netorders"),
             ("Net Order $", "netsale$")]

PREFIX = re.compile(r"^@\w+ - ")


def to_off(d):
    return (d - BASE).days


def off_to_date(off):
    return BASE + datetime.timedelta(days=off)


def date_of(v):
    if v is None or v in ("", "-"):
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return datetime.date.fromisoformat(str(v)[:10])


def num(v):
    if v is None or v in ("-", ""):
        return 0.0
    return float(v)


def load_roster():
    roster = []
    with open(ROSTER_PATH, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            team = row["Team"].strip()
            name = row["Agent"].strip()
            start = row["StartDate"].strip() or None
            end = row["EndDate"].strip() or None
            if not team or not name:
                continue
            roster.append((team, name, start, end))
    if not roster:
        sys.exit(f"No roster entries found in {ROSTER_PATH}")
    return roster


def build_tenures(roster):
    """name(lower) -> [(start_off|None, end_off|None, team, display_name)] sorted by start"""
    tenures = defaultdict(list)
    for team, name, start, end in roster:
        s = to_off(datetime.date.fromisoformat(start)) if start else None
        e = to_off(datetime.date.fromisoformat(end)) if end else None
        tenures[name.lower()].append((s, e, team, name))
    for stints in tenures.values():
        stints.sort(key=lambda t: (t[0] is None, t[0]))
    return tenures


def read_export(path, tenures):
    """Read one export, attribute rows to roster tenures, aggregate.

    Returns None if the file doesn't have the required columns, else a dict:
      agg:   (name_lc, team, store, cal_off, set_off) -> [11 metric sums]
      display: name_lc -> display name
      plus row stats and calendar-date coverage.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return None
    ws = wb.active
    it = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    hdr = [str(h).strip() if h is not None else "" for h in next(it)]
    col = {h: i for i, h in enumerate(hdr)}
    if any(c not in col for c in COLS):
        wb.close()
        return None
    idx = [col[c] for c in COLS]

    agg = defaultdict(lambda: [0.0] * len(METRICS))
    display = {}
    rows = used = out_of_tenure = not_on_roster = 0
    mind = maxd = None
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = [raw[i] for i in idx]
        rows += 1
        cal = date_of(row[0])
        if cal is None:
            continue
        if mind is None or cal < mind:
            mind = cal
        if maxd is None or cal > maxd:
            maxd = cal
        name = PREFIX.sub("", str(row[1])).strip()
        stints = tenures.get(name.lower())
        if not stints:
            not_on_roster += 1
            continue
        c = to_off(cal)
        team = None
        for s, e, t, disp in stints:
            if (s is None or c >= s) and (e is None or c <= e):
                team, name = t, disp
                break
        if team is None:
            out_of_tenure += 1
            continue
        sd = date_of(row[14])
        soff = to_off(sd) if sd is not None else c
        acc = agg[(name.lower(), team, str(row[2]).strip(), c, soff)]
        mets = [num(x) for x in row[3:14]]
        for i in range(len(METRICS)):
            acc[i] += mets[i]
        used += 1
        display[name.lower()] = name
    wb.close()
    if mind is None or not agg:
        return None
    return {"path": path, "agg": dict(agg), "display": display,
            "rows": rows, "used": used, "out_of_tenure": out_of_tenure,
            "not_on_roster": not_on_roster,
            "min_off": to_off(mind), "max_off": to_off(maxd),
            "mtime": os.path.getmtime(path)}


def merge(exports):
    """Newest export wins each (calendar day, rep+team) it covers; older files fill gaps.

    An export 'covers' a (day, rep+team) if the day is inside its calendar-date range
    and that rep+team pair appears anywhere in the file.
    """
    claim = {}
    for ex in sorted(exports, key=lambda e: -e["mtime"]):
        pairs = {(k[0], k[1]) for k in ex["agg"]}
        for off in range(ex["min_off"], ex["max_off"] + 1):
            for p in pairs:
                claim.setdefault((off, p), ex["path"])

    merged = defaultdict(lambda: [0.0] * len(METRICS))
    display = {}
    contributed = defaultdict(int)
    for ex in exports:
        display.update(ex["display"])
        for k, v in ex["agg"].items():
            if claim.get((k[3], (k[0], k[1]))) == ex["path"]:
                acc = merged[k]
                for i in range(len(METRICS)):
                    acc[i] += v[i]
                contributed[ex["path"]] += 1
    return merged, display, contributed


def build_master(merged, display):
    agents = sorted({(display[k[0]], k[1]) for k in merged}, key=lambda a: (a[1], a[0]))
    a_idx = {a: i for i, a in enumerate(agents)}
    stores = sorted({k[2] for k in merged})
    s_idx = {s: i for i, s in enumerate(stores)}

    def clean(v):
        r = round(v, 2)
        return int(r) if float(r).is_integer() else r

    recs = []
    for (name_lc, team, store, c, soff), vals in merged.items():
        recs.append([a_idx[(display[name_lc], team)], s_idx[store], c, soff]
                    + [clean(v) for v in vals])
    recs.sort(key=lambda r: (r[2], r[0], r[1], r[3]))

    return {
        "agents": [a[0] for a in agents],
        "stores": stores,
        "teams": [a[1] for a in agents],
        "metrics": METRICS,
        "nAct": N_ACT,
        "base": BASE.isoformat(),
        "recs": recs,
    }


def read_cc_export(path):
    """Parse a cost-center store export. Returns {store: [metrics]} or None if not one."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return None
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        hdr = [str(h).strip() if h is not None else "" for h in next(it)]
    except StopIteration:
        wb.close()
        return None
    col = {h: i for i, h in enumerate(hdr)}
    if "StoreName" not in col or "CalendarDate" in col or "ApptSetBy" in col \
            or any(c not in col for c, _ in CC_COLMAP):
        wb.close()
        return None
    idx = [col[c] for c, _ in CC_COLMAP]
    si = col["StoreName"]

    def cnum(v):
        if v in (None, "-", ""):
            return 0.0
        r = round(float(v), 2)
        return int(r) if float(r).is_integer() else r

    stores = {}
    for row in it:
        name = str(row[si]).strip()
        if not name or name.lower() in ("totals", "none"):
            continue
        stores[name] = [cnum(row[i]) for i in idx]
    wb.close()
    return stores or None


def js_safe(payload):
    """Serialize for embedding inside the dashboard's <script> block.

    json.dumps leaves < > & literal, so a value containing e.g. "</script>" would
    terminate the script element and inject markup into the page. Escape those
    characters (plus the JS line separators) as \\uXXXX, which decode back to the
    original characters when the browser parses the string literals.
    """
    j = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
    return (j.replace("&", "\\u0026").replace("<", "\\u003c").replace(">", "\\u003e")
             .replace("\u2028", "\\u2028").replace("\u2029", "\\u2029"))


def inject_cc(html, stores, period):
    cc = {"period": period, "metrics": [m for _, m in CC_COLMAP], "stores": stores}
    line = "const CC = " + js_safe(cc) + ";"
    if re.search(r"const CC = \{.*?\};", html, re.S):
        return re.sub(r"const CC = \{.*?\};", lambda _: line, html, count=1, flags=re.S)
    html, n = re.subn(r"(const ROSTER = \[.*?\];\n)", lambda m: m.group(1) + line + "\n",
                      html, count=1, flags=re.S)
    if n != 1:
        print("WARNING: could not insert cost-center data (ROSTER anchor not found).")
    return html


def extract_current_master(html):
    m = re.search(r"const MASTER = (\{.*?\});", html, re.S)
    return json.loads(m.group(1)) if m else None


def master_as_export(cur):
    """Wrap the dashboard's current embedded data as the oldest pseudo-export."""
    agg = {}
    display = {}
    ag, tm, st = cur["agents"], cur["teams"], cur["stores"]
    offs = [rec[2] for rec in cur["recs"]]
    for rec in cur["recs"]:
        name = ag[rec[0]]
        display[name.lower()] = name
        key = (name.lower(), tm[rec[0]], st[rec[1]], rec[2], rec[3])
        agg[key] = [float(v) for v in rec[4:4 + len(METRICS)]]
    return {"path": "(current dashboard data)", "agg": agg, "display": display,
            "rows": len(cur["recs"]), "used": len(cur["recs"]),
            "out_of_tenure": 0, "not_on_roster": 0,
            "min_off": min(offs), "max_off": max(offs), "mtime": 0.0}


def totals_of_master(cur):
    tot = [0.0] * len(METRICS)
    for rec in cur["recs"]:
        for i in range(len(METRICS)):
            tot[i] += rec[4 + i]
    return dict(zip(cur["metrics"], tot))


def totals_of(master):
    tot = [0.0] * len(METRICS)
    for rec in master["recs"]:
        for i in range(len(METRICS)):
            tot[i] += rec[4 + i]
    return dict(zip(METRICS, tot))


def inject(html, master, roster, stamp, data_through):
    mj = js_safe(master)
    rj = js_safe([list(r) for r in roster])
    html, n1 = re.subn(r"const MASTER = \{.*?\};", lambda _: "const MASTER = " + mj + ";",
                       html, count=1, flags=re.S)
    html, n2 = re.subn(r"const ROSTER = \[.*?\];", lambda _: "const ROSTER = " + rj + ";",
                       html, count=1, flags=re.S)
    html, n3 = re.subn(r'<div class="refreshdate">Data refreshed: .*?</div>',
                       f'<div class="refreshdate">Data refreshed: {stamp} &middot; '
                       f'Data through: {data_through}</div>', html, count=1)
    if n1 != 1 or n2 != 1:
        sys.exit("Could not locate the MASTER/ROSTER blocks in the HTML - aborting, nothing written.")
    if n3 != 1:
        print("WARNING: refresh-date line not found; timestamp not updated.")
    return html


def main():
    ap = argparse.ArgumentParser(description="Rebuild the dashboard's embedded data from Qlik exports.")
    ap.add_argument("--files", nargs="+", help="explicit export .xlsx paths (skips Downloads discovery)")
    ap.add_argument("--lookback", type=float, default=24,
                    help="when discovering in Downloads, only consider files modified in the last N hours (default 24)")
    ap.add_argument("--check", action="store_true", help="build and report, but do not write the HTML")
    ap.add_argument("--no-carry", action="store_true",
                    help="do NOT carry forward the dashboard's current data; rebuild strictly from the exports")
    args = ap.parse_args()

    roster = load_roster()
    tenures = build_tenures(roster)
    print(f"Roster: {len(roster)} tenure entries, {len(tenures)} distinct agents")

    if args.files:
        paths = args.files
    else:
        cutoff = datetime.datetime.now().timestamp() - args.lookback * 3600
        paths = [p for p in glob.glob(os.path.join(DOWNLOADS, "*.xlsx"))
                 if os.path.getmtime(p) >= cutoff]
        print(f"Scanning {len(paths)} recent .xlsx files in Downloads "
              f"(last {args.lookback:g}h) for the 23-column detail export ...")

    exports = []
    cc_found = None  # newest cost-center store export: (mtime, path, stores)
    for p in paths:
        ex = read_export(p, tenures)
        if ex is not None:
            exports.append(ex)
            continue
        cc = read_cc_export(p)
        if cc is not None:
            mt = os.path.getmtime(p)
            if cc_found is None or mt > cc_found[0]:
                cc_found = (mt, p, cc)
            continue
        if args.files:
            sys.exit(f"{p}: not a detail export (23 columns) or a cost-center store export")

    if not exports:
        sys.exit("No matching export files found. Export the detail table from Qlik first, "
                 "or pass --files <path> ...")

    # Drop exact duplicate downloads (same size + coverage + row count), keep newest.
    seen = {}
    for ex in sorted(exports, key=lambda e: -e["mtime"]):
        key = (os.path.getsize(ex["path"]), ex["min_off"], ex["max_off"], ex["rows"])
        if key not in seen:
            seen[key] = ex
    dropped_dupes = len(exports) - len(seen)
    exports = list(seen.values())
    if dropped_dupes:
        print(f"(ignored {dropped_dupes} duplicate download(s))")

    with open(HTML_PATH, encoding="utf-8") as f:
        html = f.read()
    current = extract_current_master(html)

    if current and not args.no_carry:
        exports.append(master_as_export(current))
        print("Carry-forward: current dashboard data participates as the oldest source "
              "(fresh exports supersede it per day+rep; use --no-carry to disable).")

    merged, display, contributed = merge(exports)
    print("\nExports used (newest wins each day it covers a rep+team):")
    for ex in sorted(exports, key=lambda e: (e["min_off"], e["mtime"])):
        print(f"  {os.path.basename(ex['path'])}: {ex['rows']:,} rows, "
              f"{off_to_date(ex['min_off'])}..{off_to_date(ex['max_off'])}, "
              f"contributed {contributed.get(ex['path'], 0):,} record groups")

    oot = sum(e["out_of_tenure"] for e in exports)
    nor = sum(e["not_on_roster"] for e in exports)
    print(f"\nRows outside any tenure window (dropped): {oot:,}"
          f" | rows for reps not on roster (dropped): {nor:,}")

    master = build_master(merged, display)
    print(f"Result: {len(master['recs']):,} records, {len(master['agents'])} agent-team pairs, "
          f"{len(master['stores'])} stores, "
          f"dates {off_to_date(master['recs'][0][2])}..{off_to_date(master['recs'][-1][2])}")

    new_tot = totals_of(master)
    prev = totals_of_master(current) if current else None
    print(f"\n{'metric':10s} {'previous':>14s} {'new':>14s} {'change':>8s}")
    for k in METRICS:
        nv = new_tot[k]
        if prev:
            pv = prev[k]
            chg = f"{(nv - pv) / pv * 100:+.1f}%" if pv else "-"
            print(f"{k:10s} {pv:14,.0f} {nv:14,.0f} {chg:>8s}")
        else:
            print(f"{k:10s} {'?':>14s} {nv:14,.0f}")

    if args.check:
        print("\n--check: dashboard NOT modified.")
        return

    now = datetime.datetime.now()
    win = os.name == "nt"
    stamp = now.strftime("%B %#d, %Y at %#I:%M %p" if win else "%B %-d, %Y at %-I:%M %p")
    max_date = off_to_date(master["recs"][-1][2])
    data_through = max_date.strftime("%B %#d, %Y" if win else "%B %-d, %Y")
    html = inject(html, master, roster, stamp, data_through)
    if cc_found:
        period = {"start": f"{max_date.year}-01-01", "end": max_date.isoformat()}
        html = inject_cc(html, cc_found[2], period)
        print(f"Cost-center store data updated from {os.path.basename(cc_found[1])} "
              f"({len(cc_found[2])} stores, period {period['start']}..{period['end']}).")
    else:
        print("No cost-center store export found - keeping the existing cost-center card data.")
    with open(HTML_PATH, "w", encoding="utf-8", newline="\n") as f:
        f.write(html)
    print(f"\nDashboard updated: {HTML_PATH}")
    print(f"Stamped: Data refreshed: {stamp} | Data through: {data_through}")
    print("Review it in a browser, then commit the change in GitHub Desktop.")


if __name__ == "__main__":
    main()
